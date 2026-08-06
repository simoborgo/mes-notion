#!/usr/bin/env python3
"""
Prepara l'anagrafica Ferramenta/Collanti/Materiale vario di consumo per l'import in Postgres
(Fase 4 anagrafica). Legge i 3 fogli categoria + il foglio inventario fisico dal file OS1 e
scrive un CSV pronto per scripts/importa-anagrafica-ferramenta.mjs.

Deduplica per codice_os1 tenendo l'ULTIMA occorrenza nel foglio (decisione presa con l'utente:
i doppioni sono sempre lo stesso articolo/fornitore con due "Cod. fornitore" diversi).
giacenza_attuale = quantità fisica contata al 30/6 se l'articolo compare in
"INVENTARIO MP 30.06.2026", altrimenti 0.

Uso: python3 scripts/prepara-anagrafica-ferramenta.py "INVENTARIO 30.06.2026 - ....xlsx" output.csv
"""
import csv
import sys

import openpyxl

FOGLI_CATEGORIA = ["Collanti", "Ferramenta", "Materiale vario di consumo"]

COLONNE_OUT = [
    "codice_os1", "descrizione", "unita_misura", "codice_fornitore", "fornitore_nome_os1",
    "descrizione_categoria", "categoria_merceologica", "cod_inv", "inventariato",
    "prezzo_ultimo_acquisto", "giacenza_attuale",
]


def main(xlsx_path: str, out_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    inventario_mp = {}
    ws_inv = wb["INVENTARIO MP 30.06.2026"]
    for r in range(5, ws_inv.max_row + 1):
        id_prodotto = ws_inv.cell(r, 3).value
        quantita = ws_inv.cell(r, 8).value
        if id_prodotto:
            inventario_mp[str(id_prodotto).strip()] = quantita or 0

    per_codice = {}
    letti = 0
    for nome_foglio in FOGLI_CATEGORIA:
        ws = wb[nome_foglio]
        for r in range(2, ws.max_row + 1):
            codice = ws.cell(r, 1).value
            if not codice:
                continue
            letti += 1
            codice = str(codice).strip()
            prezzo = ws.cell(r, 16).value
            per_codice[codice] = {
                "codice_os1": codice,
                "descrizione": (ws.cell(r, 2).value or "").strip(),
                "unita_misura": (ws.cell(r, 3).value or "").strip(),
                "codice_fornitore": (ws.cell(r, 5).value or "") if ws.cell(r, 5).value else "",
                "fornitore_nome_os1": (ws.cell(r, 6).value or "") if ws.cell(r, 6).value else "",
                "descrizione_categoria": (ws.cell(r, 7).value or "").strip(),
                "categoria_merceologica": (ws.cell(r, 10).value or "").strip(),
                "cod_inv": (ws.cell(r, 12).value or "").strip(),
                "inventariato": "true" if ws.cell(r, 15).value == "SI" else "false",
                "prezzo_ultimo_acquisto": prezzo if prezzo not in (None, "") else "",
                "giacenza_attuale": inventario_mp.get(codice, 0),
            }

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLONNE_OUT)
        writer.writeheader()
        for row in per_codice.values():
            writer.writerow(row)

    print(f"Righe lette (con duplicati): {letti}")
    print(f"Codici articolo unici scritti: {len(per_codice)}")
    print(f"Di cui con giacenza da inventario fisico: {sum(1 for v in per_codice.values() if v['giacenza_attuale'])}")
    print(f"Di cui inventariato=true: {sum(1 for v in per_codice.values() if v['inventariato'] == 'true')}")
    print(f"CSV scritto in: {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python3 scripts/prepara-anagrafica-ferramenta.py <file.xlsx> <output.csv>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
